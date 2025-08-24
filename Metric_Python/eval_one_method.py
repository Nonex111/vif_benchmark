"""
单一融合方法批量图像质量评价脚本。

本程序用于对某一融合方法在整个数据集上的融合结果进行多项主流评价指标的批量计算，并将结果保存到Excel文件。

用法：
    直接运行本脚本，将会遍历指定数据集下所有图像，对每一对红外、可见光及融合图像进行评价，并将所有结果及均值、标准差写入Excel。
    可根据需要修改 __main__ 部分的路径和方法名。

依赖：
    - PIL
    - numpy
    - openpyxl
    - tqdm
    - natsort
    - Metric.py（需包含各评价函数）

示例：
    python eval_one_method.py
"""
import numpy as np
from PIL import Image
from Metric import *
from natsort import natsorted
from tqdm import tqdm
import os
import statistics
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")  # 忽略所有警告信息，保持输出整洁

# 写入Excel的函数
def write_excel(excel_name='metric.xlsx', worksheet_name='VIF', column_index=0, data=None):
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
    # 文件不存在，创建新的 Workbook
        workbook = Workbook()

    # 获取或创建一个工作表
    if worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
    else:
        worksheet = workbook.create_sheet(title=worksheet_name)

    # 在指定列中插入数据
    column = get_column_letter(column_index + 1)
    for i, value in enumerate(data):
        cell = worksheet[column + str(i+1)]
        cell.value = value
        
    # 保存文件
    workbook.save(excel_name)

# 评估单张图像的各项指标的函数
def evaluation_one(ir_name, vi_name, f_name):
    # 读取融合图像和源图像，并转换为灰度图
    f_img = Image.open(f_name).convert('L')
    ir_img = Image.open(ir_name).convert('L')
    vi_img = Image.open(vi_name).convert('L')

    f_img_int = np.array(f_img).astype(np.int32)  # 融合图像转为int32
    f_img_double = np.array(f_img).astype(np.float32)  # 融合图像转为float32

    ir_img_int = np.array(ir_img).astype(np.int32)     # 红外图像转为int32
    ir_img_double = np.array(ir_img).astype(np.float32) # 红外图像转为float32

    vi_img_int = np.array(vi_img).astype(np.int32)     # 可见光图像转为int32
    vi_img_double = np.array(vi_img).astype(np.float32) # 可见光图像转为float32

    # 计算各项评价指标
    EN = EN_function(f_img_int)  # 熵
    MI = MI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256)  # 互信息

    SF = SF_function(f_img_double)  # 空间频率
    SD = SD_function(f_img_double)  # 标准差
    AG = AG_function(f_img_double)  # 平均梯度
    PSNR = PSNR_function(ir_img_double, vi_img_double, f_img_double)  # 峰值信噪比
    MSE = MSE_function(ir_img_double, vi_img_double, f_img_double)    # 均方误差
    VIF = VIF_function(ir_img_double, vi_img_double, f_img_double)    # 视觉信息保真度
    CC = CC_function(ir_img_double, vi_img_double, f_img_double)      # 相关系数
    SCD = SCD_function(ir_img_double, vi_img_double, f_img_double)    # 结构内容差异
    Qabf = Qabf_function(ir_img_double, vi_img_double, f_img_double)  # Qabf指标
    Nabf = Nabf_function(ir_img_double, vi_img_double, f_img_double)  # Nabf指标
    SSIM = SSIM_function(ir_img_double, vi_img_double, f_img_double)  # 结构相似性
    MS_SSIM = MS_SSIM_function(ir_img_double, vi_img_double, f_img_double)  # 多尺度结构相似性
    return EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM

if __name__ == '__main__':
    with_mean = True  # 是否计算均值和标准差
    EN_list = []      # 熵
    MI_list = []      # 互信息
    SF_list = []      # 空间频率
    AG_list = []      # 平均梯度
    SD_list = []      # 标准差
    CC_list = []      # 相关系数
    SCD_list = []     # 结构内容差异
    VIF_list = []     # 视觉信息保真度
    MSE_list = []     # 均方误差
    PSNR_list = []    # 峰值信噪比
    Qabf_list = []    # Qabf指标
    Nabf_list = []    # Nabf指标
    SSIM_list = []    # 结构相似性
    MS_SSIM_list = [] # 多尺度结构相似性
    filename_list = ['']  # 文件名列表
    dataset_name = 'test_imgs'  # 数据集名称
    ir_dir = os.path.join('..\datasets', dataset_name, 'ir')  # 红外图像文件夹
    vi_dir = os.path.join('..\datasets', dataset_name, 'vi')  # 可见光图像文件夹
    Method = 'SeAFusion'  # 融合方法名称
    f_dir = os.path.join('..\Results', dataset_name, Method)  # 融合结果文件夹
    save_dir = '..\Metric'  # 指标保存文件夹
    os.makedirs(save_dir, exist_ok=True)  # 创建保存文件夹
    metric_save_name = os.path.join(save_dir, 'metric_{}_{}.xlsx'.format(dataset_name, Method))  # Excel文件名
    filelist = natsorted(os.listdir(ir_dir))  # 获取红外图像文件名列表
    eval_bar = tqdm(filelist)  # 进度条
    for _, item in enumerate(eval_bar):
        ir_name = os.path.join(ir_dir, item)  # 红外图像路径
        vi_name = os.path.join(vi_dir, item)  # 可见光图像路径
        f_name = os.path.join(f_dir, item)    # 融合图像路径
        EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM = evaluation_one(ir_name, vi_name, f_name)  # 计算各项指标
        EN_list.append(EN)
        MI_list.append(MI)
        SF_list.append(SF)
        AG_list.append(AG)
        SD_list.append(SD)
        CC_list.append(CC)
        SCD_list.append(SCD)
        VIF_list.append(VIF)
        MSE_list.append(MSE)
        PSNR_list.append(PSNR)
        Qabf_list.append(Qabf)
        Nabf_list.append(Nabf)
        SSIM_list.append(SSIM)
        MS_SSIM_list.append(MS_SSIM)
        filename_list.append(item)
        eval_bar.set_description("{} | {}".format(Method, item))  # 更新进度条描述
    if with_mean:
    # 添加均值
        EN_list.append(np.mean(EN_list))
        MI_list.append(np.mean(MI_list))
        SF_list.append(np.mean(SF_list))
        AG_list.append(np.mean(AG_list))
        SD_list.append(np.mean(SD_list))
        CC_list.append(np.mean(CC_list))
        SCD_list.append(np.mean(SCD_list))
        VIF_list.append(np.mean(VIF_list))
        MSE_list.append(np.mean(MSE_list))
        PSNR_list.append(np.mean(PSNR_list))
        Qabf_list.append(np.mean(Qabf_list))
        Nabf_list.append(np.mean(Nabf_list))
        SSIM_list.append(np.mean(SSIM_list))
        MS_SSIM_list.append(np.mean(MS_SSIM_list))
        filename_list.append('mean')

        ## 添加标准差
        EN_list.append(np.std(EN_list))
        MI_list.append(np.std(MI_list))
        SF_list.append(np.std(SF_list))
        AG_list.append(np.std(AG_list))
        SD_list.append(np.std(SD_list))
        CC_list.append(np.std(CC_list[:-1]))
        SCD_list.append(np.std(SCD_list))
        VIF_list.append(np.std(VIF_list))
        MSE_list.append(np.std(MSE_list))
        PSNR_list.append(np.std(PSNR_list))
        Qabf_list.append(np.std(Qabf_list))
        Nabf_list.append(np.std(Nabf_list))
        SSIM_list.append(np.std(SSIM_list))
        MS_SSIM_list.append(np.std(MS_SSIM_list))
        filename_list.append('std')

    ## 保留三位小数
    EN_list = [round(x, 3) for x in EN_list]
    MI_list = [round(x, 3) for x in MI_list]
    SF_list = [round(x, 3) for x in SF_list]
    AG_list = [round(x, 3) for x in AG_list]
    SD_list = [round(x, 3) for x in SD_list]
    CC_list = [round(x, 3) for x in CC_list]
    SCD_list = [round(x, 3) for x in SCD_list]
    VIF_list = [round(x, 3) for x in VIF_list]
    MSE_list = [round(x, 3) for x in MSE_list]
    PSNR_list = [round(x, 3) for x in PSNR_list]
    Qabf_list = [round(x, 3) for x in Qabf_list]
    Nabf_list = [round(x, 3) for x in Nabf_list]
    SSIM_list = [round(x, 3) for x in SSIM_list]
    MS_SSIM_list = [round(x, 3) for x in MS_SSIM_list]

    EN_list.insert(0, '{}'.format(Method))
    MI_list.insert(0, '{}'.format(Method))
    SF_list.insert(0, '{}'.format(Method))
    AG_list.insert(0, '{}'.format(Method))
    SD_list.insert(0, '{}'.format(Method))
    CC_list.insert(0, '{}'.format(Method))
    SCD_list.insert(0, '{}'.format(Method))
    VIF_list.insert(0, '{}'.format(Method))
    MSE_list.insert(0, '{}'.format(Method))
    PSNR_list.insert(0, '{}'.format(Method))
    Qabf_list.insert(0, '{}'.format(Method))
    Nabf_list.insert(0, '{}'.format(Method))
    SSIM_list.insert(0, '{}'.format(Method))
    MS_SSIM_list.insert(0, '{}'.format(Method))
    write_excel(metric_save_name, 'EN', 0, filename_list)
    write_excel(metric_save_name, "MI", 0, filename_list)
    write_excel(metric_save_name, "SF", 0, filename_list)
    write_excel(metric_save_name, "AG", 0, filename_list)
    write_excel(metric_save_name, "SD", 0, filename_list)
    write_excel(metric_save_name, "CC", 0, filename_list)
    write_excel(metric_save_name, "SCD", 0, filename_list)
    write_excel(metric_save_name, "VIF", 0, filename_list)
    write_excel(metric_save_name, "MSE", 0, filename_list)
    write_excel(metric_save_name, "PSNR", 0, filename_list)
    write_excel(metric_save_name, "Qabf", 0, filename_list)
    write_excel(metric_save_name, "Nabf", 0, filename_list)
    write_excel(metric_save_name, "SSIM", 0, filename_list)
    write_excel(metric_save_name, "MS_SSIM", 0, filename_list)
    write_excel(metric_save_name, 'EN', 1, EN_list)
    write_excel(metric_save_name, 'MI', 1, MI_list)
    write_excel(metric_save_name, 'SF', 1, SF_list)
    write_excel(metric_save_name, 'AG', 1, AG_list)
    write_excel(metric_save_name, 'SD', 1, SD_list)
    write_excel(metric_save_name, 'CC', 1, CC_list)
    write_excel(metric_save_name, 'SCD', 1, SCD_list)
    write_excel(metric_save_name, 'VIF', 1, VIF_list)
    write_excel(metric_save_name, 'MSE', 1, MSE_list)
    write_excel(metric_save_name, 'PSNR', 1, PSNR_list)
    write_excel(metric_save_name, 'Qabf', 1, Qabf_list)
    write_excel(metric_save_name, 'Nabf', 1, Nabf_list)
    write_excel(metric_save_name, 'SSIM', 1, SSIM_list)
    write_excel(metric_save_name, 'MS_SSIM', 1, MS_SSIM_list)